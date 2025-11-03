import json
import csv
from pathlib import Path


# Реализация функций конвертации
def json_to_csv(json_path: str, csv_path: str) -> None:
    """
    Преобразует JSON-файл в CSV с валидацией.
    """
    # Проверка существования файла
    json_file = Path(json_path)
    if not json_file.exists():
        raise FileNotFoundError(f"JSON файл не найден: {json_path}")

    # Чтение JSON
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise ValueError(f"Ошибка декодирования JSON: {e}")

    # Валидация структуры JSON
    if not isinstance(data, list):
        raise ValueError("JSON должен содержать список объектов")

    if not data:
        raise ValueError("Пустой JSON или неподдерживаемая структура")

    if not all(isinstance(item, dict) for item in data):
        raise ValueError("Все элементы JSON должны быть словарями")

    # Определение колонок
    all_keys = set()
    for item in data:
        all_keys.update(item.keys())

    first_item_keys = list(data[0].keys()) if data else []
    remaining_keys = sorted(all_keys - set(first_item_keys))
    fieldnames = first_item_keys + remaining_keys

    # Запись CSV
    Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row in data:
            complete_row = {key: str(row.get(key, '')) for key in fieldnames}
            writer.writerow(complete_row)


def csv_to_json(csv_path: str, json_path: str) -> None:
    """
    Преобразует CSV в JSON с валидацией.
    """
    # Проверка существования файла
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    # Чтение CSV
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            # Проверка заголовка
            if reader.fieldnames is None:
                raise ValueError("CSV файл не содержит заголовка")

            data = list(reader)

    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")

    # Валидация CSV
    if not data:
        raise ValueError("CSV файл пуст")

    # Запись JSON
    Path(json_path).parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX с использованием openpyxl.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для работы функции требуется установить openpyxl: pip install openpyxl")

    # Проверка существования файла
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    # Чтение CSV
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            data = list(reader)
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")

    # Проверка на пустой CSV
    if not data:
        raise ValueError("CSV файл пуст")

    # Создание XLSX
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Запись данных в лист
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Настройка автоширины колонок (не менее 8 символов)
    for col_idx, _ in enumerate(data[0], 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0

        # Находим максимальную длину в колонке
        for row in ws[column_letter]:
            if row.value:
                max_length = max(max_length, len(str(row.value)))

        # Устанавливаем ширину (не менее 8 символов)
        adjusted_width = max(max_length + 2, 8)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохранение файла
    wb.save(xlsx_path)


# Создание примеров файлов
def create_sample_files():
    """Создает примеры файлов для демонстрации"""
    Path("data/samples").mkdir(parents=True, exist_ok=True)
    Path("data/out").mkdir(parents=True, exist_ok=True)

    # 1. people.json - список объектов одинаковой формы
    people_data = [
        {"name": "Анна", "age": 28, "city": "Москва", "occupation": "Инженер"},
        {"name": "Петр", "age": 35, "city": "СПБ", "occupation": "Дизайнер"},
        {"name": "Мария", "age": 24, "city": "Казань", "occupation": "Аналитик"}
    ]

    with open("data/samples/people.json", "w", encoding="utf-8") as f:
        json.dump(people_data, f, ensure_ascii=False, indent=2)

    # 2. people.csv - те же данные для обратимого преобразования
    with open("data/samples/people.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "age", "city", "occupation"])
        writer.writeheader()
        writer.writerows(people_data)

    # 3. cities.csv - для демонстрации CSV→XLSX
    cities_data = [
        {"city": "Москва", "population": "12678079", "area": "2561", "foundation": "1147"},
        {"city": "Санкт-Петербург", "population": "5384342", "area": "1439", "foundation": "1703"},
        {"city": "Новосибирск", "population": "1625631", "area": "505", "foundation": "1893"},
        {"city": "Екатеринбург", "population": "1493749", "area": "468", "foundation": "1723"}
    ]

    with open("data/samples/cities.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["city", "population", "area", "foundation"])
        writer.writeheader()
        writer.writerows(cities_data)

    print("Примеры файлов созданы:")
    print("  - data/samples/people.json")
    print("  - data/samples/people.csv")
    print("  - data/samples/cities.csv")


# Демонстрационные сценарии
def demo_json_to_csv():
    """Сценарий 1: JSON→CSV с проверкой"""
    print("\n=== СЦЕНАРИЙ 1: JSON → CSV ===")

    # Преобразование
    json_to_csv("data/samples/people.json", "data/out/people_from_json.csv")

    # Проверка
    with open("data/samples/people.json", encoding="utf-8") as f:
        json_data = json.load(f)

    with open("data/out/people_from_json.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        csv_data = list(reader)

    print(f" Исходный JSON: {len(json_data)} записей")
    print(f" Полученный CSV: {len(csv_data)} записей")
    print(f" Заголовок CSV: {', '.join(csv_data[0].keys())}")
    print(f" Соответствие записей: {len(json_data) == len(csv_data)}")


def demo_csv_to_json():
    """Сценарий 2: CSV→JSON со сравнением"""
    print("\n=== СЦЕНАРИЙ 2: CSV → JSON ===")

    # Преобразование
    csv_to_json("data/samples/people.csv", "data/out/people_from_csv.json")

    # Проверка
    with open("data/samples/people.csv", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        original_csv = list(reader)

    with open("data/out/people_from_csv.json", encoding="utf-8") as f:
        converted_json = json.load(f)

    print(f" Исходный CSV: {len(original_csv)} записей")
    print(f" Полученный JSON: {len(converted_json)} записей")
    print(f" Ключи в JSON: {list(converted_json[0].keys())}")
    print(f" Соответствие структуры: {len(original_csv) == len(converted_json)}")


def demo_csv_to_xlsx():
    """Сценарий 3: CSV→XLSX с автошириной"""
    print("\n=== СЦЕНАРИЙ 3: CSV → XLSX ===")

    # Преобразование
    csv_to_xlsx("data/samples/cities.csv", "data/out/cities.xlsx")

    # Проверка
    with open("data/samples/cities.csv", encoding="utf-8") as f:
        reader = csv.reader(f)
        csv_rows = list(reader)

    print(f" Исходный CSV: {len(csv_rows)} строк, {len(csv_rows[0])} колонок")
    print(f" Создан XLSX: data/out/cities.xlsx")
    print(" Проверьте в Excel/LibreOffice:")
    print("  - Лист 'Sheet1' с данными")
    print("  - Автоширина колонок (не менее 8 символов)")
    print("  - Корректное отображение кириллицы")


# Запуск полной демонстрации
if __name__ == "__main__":
    print("Создание примеров файлов...")
    create_sample_files()

    print("\nЗапуск демонстрационных сценариев...")
    demo_json_to_csv()
    demo_csv_to_json()
    demo_csv_to_xlsx()

    print("\n Все сценарии завершены успешно!")