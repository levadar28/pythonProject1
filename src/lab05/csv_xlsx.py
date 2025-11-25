import csv
from pathlib import Path

# Проверяем наличие openpyxl и импортируем
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    # Проверка доступности openpyxl
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "Для работы функции требуется установить openpyxl: pip install openpyxl"
        )

    # Проверка существования файла
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    # Проверка расширения файла
    if csv_file.suffix.lower() != ".csv":
        raise ValueError(
            f"Неверный тип файла: ожидается .csv, получен {csv_file.suffix}"
        )

    # Проверка расширения выходного файла
    xlsx_file = Path(xlsx_path)
    if xlsx_file.suffix.lower() != ".xlsx":
        raise ValueError(
            f"Неверный тип выходного файла: ожидается .xlsx, получен {xlsx_file.suffix}"
        )

    # Чтение CSV
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            data = list(reader)
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")

    # Проверка на пустой CSV
    if not data:
        raise ValueError("CSV файл пуст")

    # Проверка наличия заголовка
    if not data[0]:
        raise ValueError("Первая строка CSV (заголовок) пуста")

    # Создание XLSX
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Запись данных в лист
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Настройка автоширины колонок (не менее 8 символов)
        for col_idx, _ in enumerate(data[0], 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            # Находим максимальную длину в колонке
            for row in ws[column_letter]:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))

            # Устанавливаем ширину (не менее 8 символов)
            adjusted_width = max(max_length + 2, 8)  # +2 для отступов
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохранение файла
        wb.save(xlsx_path)

    except Exception as e:
        raise ValueError(f"Ошибка создания XLSX: {e}")
